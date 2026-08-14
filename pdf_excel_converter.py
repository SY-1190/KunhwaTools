from __future__ import annotations

import io
import math
import re
from dataclasses import dataclass, field
from statistics import median
from typing import Iterable

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BBox = tuple[float, float, float, float]


@dataclass
class TextLine:
    text: str
    bbox: BBox
    font_size: float
    bold: bool
    side: int | None = None


@dataclass
class TableCell:
    bbox: BBox
    text: str
    bold: bool
    font_size: float
    fill: str | None


@dataclass
class TableBlock:
    bbox: BBox
    cells: list[TableCell]
    x_bounds: list[float]
    y_bounds: list[float]

    @property
    def rows(self) -> int:
        return len(self.y_bounds) - 1

    @property
    def columns(self) -> int:
        return len(self.x_bounds) - 1


@dataclass
class PageModel:
    number: int
    width: float
    height: float
    lines: list[TextLine]
    tables: list[TableBlock]
    column_split: float | None
    preview_png: bytes
    warnings: list[str] = field(default_factory=list)


def _area(bbox: BBox) -> float:
    return max(0.0, bbox[2] - bbox[0]) * max(0.0, bbox[3] - bbox[1])


def _contains(outer: BBox, inner: BBox, tolerance: float = 0.8) -> bool:
    return (
        outer[0] <= inner[0] + tolerance
        and outer[1] <= inner[1] + tolerance
        and outer[2] >= inner[2] - tolerance
        and outer[3] >= inner[3] - tolerance
    )


def _inside(point: tuple[float, float], bbox: BBox, tolerance: float = 0.5) -> bool:
    return (
        bbox[0] - tolerance <= point[0] <= bbox[2] + tolerance
        and bbox[1] - tolerance <= point[1] <= bbox[3] + tolerance
    )


def _overlap(start_a: float, end_a: float, start_b: float, end_b: float) -> float:
    return max(0.0, min(end_a, end_b) - max(start_a, start_b))


def _touches(left: BBox, right: BBox, tolerance: float = 1.5) -> bool:
    vertical_overlap = _overlap(left[1], left[3], right[1], right[3])
    horizontal_overlap = _overlap(left[0], left[2], right[0], right[2])
    shares_vertical = (
        abs(left[2] - right[0]) <= tolerance or abs(right[2] - left[0]) <= tolerance
    ) and vertical_overlap > tolerance
    shares_horizontal = (
        abs(left[3] - right[1]) <= tolerance or abs(right[3] - left[1]) <= tolerance
    ) and horizontal_overlap > tolerance
    return shares_vertical or shares_horizontal


def _cluster_values(values: Iterable[float], tolerance: float = 1.5) -> list[float]:
    ordered = sorted(values)
    if not ordered:
        return []
    clusters: list[list[float]] = [[ordered[0]]]
    for value in ordered[1:]:
        if abs(value - sum(clusters[-1]) / len(clusters[-1])) <= tolerance:
            clusters[-1].append(value)
        else:
            clusters.append([value])
    return [sum(cluster) / len(cluster) for cluster in clusters]


def _nearest_index(values: list[float], target: float) -> int:
    return min(range(len(values)), key=lambda index: abs(values[index] - target))


def _clean_text(text: str | None, merge_spaced_glyphs: bool = True) -> str:
    if not text:
        return ""
    lines = []
    for raw_line in text.replace("\u00a0", " ").splitlines():
        line = re.sub(r"[ \t]+", " ", raw_line).strip()
        line = re.sub(r"^[◇◆◈◊]+\s*", "", line)
        if not line:
            continue
        tokens = line.split(" ")
        if not merge_spaced_glyphs:
            lines.append(line)
            continue
        merged: list[str] = []
        index = 0
        while index < len(tokens):
            token = tokens[index]
            if len(token) == 1 and re.fullmatch(r"[0-9A-Za-z가-힣]", token):
                run = [token]
                next_index = index + 1
                while (
                    next_index < len(tokens)
                    and len(tokens[next_index]) == 1
                    and re.fullmatch(r"[0-9A-Za-z가-힣]", tokens[next_index])
                ):
                    run.append(tokens[next_index])
                    next_index += 1
                if len(run) >= 2:
                    merged.append("".join(run))
                    index = next_index
                    continue
            merged.append(token)
            index += 1
        line = " ".join(merged)
        line = re.sub(r"(?<=\d)\s+(?=[\d,])", "", line)
        line = re.sub(r"(?<=,)\s+(?=\d)", "", line)
        lines.append(line)
    return "\n".join(lines)


def _color_to_argb(color: object) -> str | None:
    if color is None:
        return None
    if isinstance(color, (int, float)):
        value = max(0, min(255, round(float(color) * 255)))
        return f"FF{value:02X}{value:02X}{value:02X}"
    if not isinstance(color, (tuple, list)):
        return None
    values = [float(item) for item in color]
    if len(values) == 1:
        value = max(0, min(255, round(values[0] * 255)))
        return f"FF{value:02X}{value:02X}{value:02X}"
    if len(values) == 3:
        rgb = [max(0, min(255, round(item * 255))) for item in values]
        return "FF" + "".join(f"{item:02X}" for item in rgb)
    if len(values) == 4:
        cyan, magenta, yellow, black = values
        rgb = [
            round(255 * (1 - min(1, cyan * (1 - black) + black))),
            round(255 * (1 - min(1, magenta * (1 - black) + black))),
            round(255 * (1 - min(1, yellow * (1 - black) + black))),
        ]
        return "FF" + "".join(f"{item:02X}" for item in rgb)
    return None


def _is_visible_text_color(color: object) -> bool:
    if color is None:
        return True
    if isinstance(color, (int, float)):
        return float(color) < 0.97
    if not isinstance(color, (tuple, list)):
        return True
    values = [float(item) for item in color]
    if len(values) == 1:
        return values[0] < 0.97
    if len(values) == 3:
        return not all(item >= 0.97 for item in values)
    if len(values) == 4:
        return not all(item <= 0.03 for item in values)
    return True


def _cell_appearance(page: pdfplumber.page.Page, bbox: BBox) -> tuple[bool, float, str | None]:
    chars = []
    for char in page.chars:
        if not _is_visible_text_color(char.get("non_stroking_color")):
            continue
        center = ((float(char["x0"]) + float(char["x1"])) / 2, (float(char["top"]) + float(char["bottom"])) / 2)
        if _inside(center, bbox, tolerance=0.2):
            chars.append(char)
    sizes = [float(char.get("size") or 10) for char in chars]
    bold = any("bold" in str(char.get("fontname") or "").lower() for char in chars)
    font_size = median(sizes) if sizes else 10.0
    center = ((bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2)
    matching_rects = []
    for rect in page.rects:
        rect_bbox = (float(rect["x0"]), float(rect["top"]), float(rect["x1"]), float(rect["bottom"]))
        if _inside(center, rect_bbox) and _area(rect_bbox) <= _area(bbox) * 1.35:
            matching_rects.append(rect)
    matching_rects.sort(key=lambda rect: float(rect["width"]) * float(rect["height"]))
    fill = _color_to_argb(matching_rects[0].get("non_stroking_color")) if matching_rects else None
    if fill in {"FFFFFFFF", "FFFEFEFE"}:
        fill = None
    return bold, font_size, fill


def _extract_cell_text(page: pdfplumber.page.Page, bbox: BBox) -> str:
    chars = []
    for char in page.chars:
        if not _is_visible_text_color(char.get("non_stroking_color")):
            continue
        center = ((float(char["x0"]) + float(char["x1"])) / 2, (float(char["top"]) + float(char["bottom"])) / 2)
        if _inside(center, bbox, tolerance=0.05):
            chars.append(char)
    chars.sort(key=lambda char: (float(char["top"]), float(char["x0"])))
    lines: list[list[dict]] = []
    for char in chars:
        center = (float(char["top"]) + float(char["bottom"])) / 2
        target = None
        for line in reversed(lines[-4:]):
            line_center = sum((float(item["top"]) + float(item["bottom"])) / 2 for item in line) / len(line)
            if abs(center - line_center) <= max(1.5, float(char.get("size") or 10) * 0.3):
                target = line
                break
        if target is None:
            lines.append([char])
        else:
            target.append(char)
    rendered = []
    for line in lines:
        line.sort(key=lambda char: float(char["x0"]))
        text = ""
        previous_x1 = None
        previous_size = 10.0
        for char in line:
            value = str(char.get("text") or "")
            x0 = float(char["x0"])
            gap = 0 if previous_x1 is None else x0 - previous_x1
            if value.isspace():
                if text and not text.endswith(" "):
                    text += " "
            else:
                if text and not text.endswith(" ") and gap > max(1.8, previous_size * 0.32):
                    text += " "
                text += value
            previous_x1 = float(char["x1"])
            previous_size = float(char.get("size") or previous_size)
        rendered.append(text)
    return _clean_text("\n".join(rendered))


def _detect_table_blocks(page: pdfplumber.page.Page) -> list[TableBlock]:
    settings = {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "snap_tolerance": 2,
        "join_tolerance": 2,
        "intersection_tolerance": 2,
        "edge_min_length": 4,
        "text_x_tolerance": 3,
        "text_y_tolerance": 3,
    }
    finder = page.debug_tablefinder(settings)
    raw_cells = [tuple(float(value) for value in cell) for cell in finder.cells]
    unique_cells: list[BBox] = []
    seen = set()
    for cell in raw_cells:
        key = tuple(round(value, 1) for value in cell)
        if key not in seen:
            unique_cells.append(cell)
            seen.add(key)

    leaf_cells = []
    for cell in unique_cells:
        contains_smaller = any(
            other != cell and _contains(cell, other) and _area(cell) > _area(other) * 1.4
            for other in unique_cells
        )
        if not contains_smaller:
            leaf_cells.append(cell)

    unvisited = set(range(len(leaf_cells)))
    components: list[list[BBox]] = []
    while unvisited:
        seed = unvisited.pop()
        component_indices = {seed}
        queue = [seed]
        while queue:
            current = queue.pop()
            neighbors = [index for index in list(unvisited) if _touches(leaf_cells[current], leaf_cells[index])]
            for neighbor in neighbors:
                unvisited.remove(neighbor)
                component_indices.add(neighbor)
                queue.append(neighbor)
        components.append([leaf_cells[index] for index in sorted(component_indices)])

    blocks = []
    for component in components:
        x_bounds = _cluster_values([value for cell in component for value in (cell[0], cell[2])])
        y_bounds = _cluster_values([value for cell in component for value in (cell[1], cell[3])])
        if len(component) < 4 or len(x_bounds) < 3 or len(y_bounds) < 3:
            continue
        bbox = (
            min(cell[0] for cell in component),
            min(cell[1] for cell in component),
            max(cell[2] for cell in component),
            max(cell[3] for cell in component),
        )
        cells = []
        nonempty = 0
        for cell_bbox in component:
            text = _extract_cell_text(page, cell_bbox)
            if text:
                nonempty += 1
            bold, font_size, fill = _cell_appearance(page, cell_bbox)
            cells.append(TableCell(cell_bbox, text, bold, font_size, fill))
        if nonempty < 2 and len(component) < 8:
            continue
        while len(x_bounds) > 2:
            segment_start = x_bounds[-2]
            if any(cell.text and cell.bbox[2] > segment_start + 0.5 for cell in cells):
                break
            cells = [cell for cell in cells if cell.bbox[0] < segment_start - 0.5]
            x_bounds.pop()
        while len(y_bounds) > 2:
            segment_start = y_bounds[-2]
            if any(cell.text and cell.bbox[3] > segment_start + 0.5 for cell in cells):
                break
            cells = [cell for cell in cells if cell.bbox[1] < segment_start - 0.5]
            y_bounds.pop()
        if len(x_bounds) < 3 or len(y_bounds) < 3:
            continue
        bbox = (x_bounds[0], y_bounds[0], x_bounds[-1], y_bounds[-1])
        blocks.append(TableBlock(bbox, cells, x_bounds, y_bounds))

    blocks.sort(key=lambda block: (block.bbox[1], block.bbox[0]))
    return blocks


def _group_word_objects(words: list[dict]) -> list[list[dict]]:
    ordered = sorted(words, key=lambda word: (float(word["top"]), float(word["x0"])))
    groups: list[list[dict]] = []
    for word in ordered:
        center = (float(word["top"]) + float(word["bottom"])) / 2
        target = None
        for group in reversed(groups[-5:]):
            group_center = sum((float(item["top"]) + float(item["bottom"])) / 2 for item in group) / len(group)
            tolerance = max(2.0, min(float(word.get("size") or 10), float(group[0].get("size") or 10)) * 0.35)
            if abs(group_center - center) <= tolerance:
                target = group
                break
        if target is None:
            groups.append([word])
        else:
            target.append(word)
    return groups


def _group_words_into_lines(words: list[dict], side: int | None = None) -> list[TextLine]:
    groups = _group_word_objects(words)

    lines = []
    for group in groups:
        group.sort(key=lambda word: float(word["x0"]))
        text = _clean_text(" ".join(str(word["text"]) for word in group), merge_spaced_glyphs=False)
        if not text:
            continue
        sizes = [float(word.get("size") or 10) for word in group]
        lines.append(
            TextLine(
                text=text,
                bbox=(
                    min(float(word["x0"]) for word in group),
                    min(float(word["top"]) for word in group),
                    max(float(word["x1"]) for word in group),
                    max(float(word["bottom"]) for word in group),
                ),
                font_size=median(sizes),
                bold=any("bold" in str(word.get("fontname") or "").lower() for word in group),
                side=side,
            )
        )
    return lines


def _detect_column_split(words: list[dict], page_width: float) -> float | None:
    if len(words) < 18:
        return None
    best: tuple[float, float] | None = None
    for step in range(35, 66, 2):
        split = page_width * step / 100
        left = [word for word in words if (float(word["x0"]) + float(word["x1"])) / 2 < split]
        right = [word for word in words if (float(word["x0"]) + float(word["x1"])) / 2 >= split]
        half_gutter = max(10.0, page_width * 0.025)
        crossing = [
            word
            for word in words
            if float(word["x0"]) < split + half_gutter and float(word["x1"]) > split - half_gutter
        ]
        if len(left) < 8 or len(right) < 8:
            continue
        crossing_ratio = len(crossing) / len(words)
        if crossing_ratio > 0.06:
            continue
        left_lines = _group_words_into_lines(left)
        right_lines = _group_words_into_lines(right)
        if len(left_lines) < 8 or len(right_lines) < 8:
            continue
        line_balance = min(len(left_lines), len(right_lines)) / max(len(left_lines), len(right_lines))
        if line_balance < 0.6:
            continue
        balance = min(len(left), len(right)) / max(len(left), len(right))
        left_span = (min(float(word["top"]) for word in left), max(float(word["bottom"]) for word in left))
        right_span = (min(float(word["top"]) for word in right), max(float(word["bottom"]) for word in right))
        overlap = _overlap(left_span[0], left_span[1], right_span[0], right_span[1])
        span_ratio = overlap / max(1.0, min(left_span[1] - left_span[0], right_span[1] - right_span[0]))
        if span_ratio < 0.35:
            continue
        score = balance * 0.35 + line_balance * 0.3 + span_ratio * 0.35 - crossing_ratio * 2.5
        if best is None or score > best[0]:
            best = (score, split)
    return best[1] if best else None


def _extract_text_lines(page: pdfplumber.page.Page, table_blocks: list[TableBlock]) -> tuple[list[TextLine], float | None]:
    words = page.extract_words(
        x_tolerance=3,
        y_tolerance=3,
        keep_blank_chars=False,
        extra_attrs=["fontname", "size", "non_stroking_color"],
    )
    words = [word for word in words if _is_visible_text_color(word.get("non_stroking_color"))]
    outside = []
    for word in words:
        center = ((float(word["x0"]) + float(word["x1"])) / 2, (float(word["top"]) + float(word["bottom"])) / 2)
        if not any(_inside(center, table.bbox, tolerance=0.2) for table in table_blocks):
            outside.append(word)

    split = None if table_blocks else _detect_column_split(outside, float(page.width))
    if split is None:
        return _group_words_into_lines(outside), None
    wide_word_ids: set[int] = set()
    for group in _group_word_objects(outside):
        ordered = sorted(group, key=lambda word: float(word["x0"]))
        group_x0 = min(float(word["x0"]) for word in ordered)
        group_x1 = max(float(word["x1"]) for word in ordered)
        left_edge = max(
            (float(word["x1"]) for word in ordered if (float(word["x0"]) + float(word["x1"])) / 2 < split),
            default=group_x0,
        )
        right_edge = min(
            (float(word["x0"]) for word in ordered if (float(word["x0"]) + float(word["x1"])) / 2 >= split),
            default=group_x1,
        )
        crosses_both_columns = group_x0 < split - float(page.width) * 0.18 and group_x1 > split + float(page.width) * 0.18
        continuous_at_gutter = right_edge - left_edge <= float(page.width) * 0.06
        if crosses_both_columns and continuous_at_gutter:
            wide_word_ids.update(id(word) for word in ordered)

    left = []
    right = []
    wide = []
    for word in outside:
        if id(word) in wide_word_ids:
            wide.append(word)
            continue
        x0 = float(word["x0"])
        x1 = float(word["x1"])
        if x0 < split < x1 and x1 - x0 > float(page.width) * 0.25:
            wide.append(word)
        elif (x0 + x1) / 2 < split:
            left.append(word)
        else:
            right.append(word)
    lines = _group_words_into_lines(left, 0) + _group_words_into_lines(right, 1) + _group_words_into_lines(wide, None)
    lines.sort(key=lambda line: (line.bbox[1], line.bbox[0]))
    return lines, split


def analyze_pdf(pdf_bytes: bytes, preview_resolution: int = 110) -> list[PageModel]:
    models = []
    with pdfplumber.open(io.BytesIO(pdf_bytes), unicode_norm="NFC") as pdf:
        for page in pdf.pages:
            tables = _detect_table_blocks(page)
            lines, split = _extract_text_lines(page, tables)
            preview = page.to_image(resolution=preview_resolution, antialias=True)
            preview_buffer = io.BytesIO()
            preview.original.save(preview_buffer, format="PNG", optimize=True)
            warnings = []
            if not page.chars:
                warnings.append("텍스트 레이어가 없어 OCR 변환이 필요합니다.")
            models.append(
                PageModel(
                    number=page.page_number,
                    width=float(page.width),
                    height=float(page.height),
                    lines=lines,
                    tables=tables,
                    column_split=split,
                    preview_png=preview_buffer.getvalue(),
                    warnings=warnings,
                )
            )
            page.close()
    return models


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name).strip()[:31] or "시트"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        marker = f"_{suffix}"
        candidate = f"{cleaned[:31-len(marker)]}{marker}"
        suffix += 1
    used.add(candidate)
    return candidate


def _is_heading(line: TextLine, median_size: float) -> bool:
    return line.bold or line.font_size >= median_size * 1.25


def _apply_cell_style(cell, font_size: float = 10, bold: bool = False, fill: str | None = None) -> None:
    cell.font = Font(name="맑은 고딕", size=max(8, min(16, font_size)), bold=bold, color="FF172033")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _write_table(ws, block: TableBlock, start_row: int, max_columns: int) -> int:
    thin = Side(style="thin", color="FF7B8494")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    local_columns = block.columns
    for column in range(local_columns):
        width_points = block.x_bounds[column + 1] - block.x_bounds[column]
        width = max(4.0, min(24.0, width_points / 6.0))
        letter = get_column_letter(column + 1)
        current = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(current, width)
    for row_index in range(block.rows):
        height_points = block.y_bounds[row_index + 1] - block.y_bounds[row_index]
        ws.row_dimensions[start_row + row_index].height = max(18, min(54, height_points * 0.95))
        for column in range(1, local_columns + 1):
            ws.cell(start_row + row_index, column).border = border

    occupied = set()
    for spec in sorted(block.cells, key=lambda item: (item.bbox[1], item.bbox[0])):
        col_start = _nearest_index(block.x_bounds, spec.bbox[0]) + 1
        col_end = _nearest_index(block.x_bounds, spec.bbox[2])
        row_start = start_row + _nearest_index(block.y_bounds, spec.bbox[1])
        row_end = start_row + _nearest_index(block.y_bounds, spec.bbox[3]) - 1
        if col_start > col_end or row_start > row_end:
            continue
        key = (row_start, col_start)
        if key in occupied:
            continue
        occupied.add(key)
        cell = ws.cell(row_start, col_start)
        cell.value = spec.text or None
        _apply_cell_style(cell, spec.font_size, spec.bold, spec.fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_end, end_column=col_end)

    if local_columns < max_columns:
        for row_index in range(start_row, start_row + block.rows):
            for column in range(local_columns + 1, max_columns + 1):
                ws.cell(row_index, column).fill = PatternFill("solid", fgColor="FFF8FAFC")
    return start_row + block.rows


def _write_page_sheet(ws, model: PageModel) -> None:
    max_columns = max((table.columns for table in model.tables), default=12)
    max_columns = max(1, max_columns)
    for column in range(1, max_columns + 1):
        ws.column_dimensions[get_column_letter(column)].width = 10 if not model.tables else 12
    font_sizes = [line.font_size for line in model.lines]
    median_size = median(font_sizes) if font_sizes else 10
    events = [(table.bbox[1], "table", table) for table in model.tables]
    events.extend((line.bbox[1], "line", line) for line in model.lines)
    events.sort(key=lambda event: (event[0], 0 if event[1] == "line" else 1))

    row = 1
    previous_bottom = 0.0
    index = 0
    while index < len(events):
        top, event_type, payload = events[index]
        if top - previous_bottom > max(20, median_size * 2.5) and row > 1:
            ws.row_dimensions[row].height = 8
            row += 1
        if event_type == "table":
            block: TableBlock = payload
            row = _write_table(ws, block, row, max_columns)
            previous_bottom = block.bbox[3]
            index += 1
            continue

        line: TextLine = payload
        paired: TextLine | None = None
        if model.column_split and line.side in (0, 1) and index + 1 < len(events):
            next_top, next_type, next_payload = events[index + 1]
            if (
                next_type == "line"
                and next_payload.side in (0, 1)
                and next_payload.side != line.side
                and abs(next_top - top) <= max(3.0, min(line.font_size, next_payload.font_size) * 0.45)
            ):
                paired = next_payload
        row_lines = [line] + ([paired] if paired else [])
        for row_line in row_lines:
            if row_line.side is None or not model.column_split:
                start_column, end_column = 1, max_columns
            else:
                left_end = max(1, max_columns // 2)
                start_column, end_column = (1, left_end) if row_line.side == 0 else (left_end + 1, max_columns)
            cell = ws.cell(row, start_column)
            cell.value = row_line.text
            heading = _is_heading(row_line, median_size)
            _apply_cell_style(cell, row_line.font_size + (1 if heading else 0), heading)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if end_column > start_column:
                ws.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)
        longest = max(len(item.text) for item in row_lines)
        available = max(20, max_columns * 12 // len(row_lines))
        line_count = max(1, math.ceil(longest / available))
        ws.row_dimensions[row].height = max(18, min(72, 15 * line_count))
        previous_bottom = max(item.bbox[3] for item in row_lines)
        row += 1
        index += 2 if paired else 1

    if model.warnings:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_columns)
        warning_cell = ws.cell(row, 1)
        warning_cell.value = " / ".join(model.warnings)
        warning_cell.font = Font(name="맑은 고딕", size=9, color="FFC2410C")
        warning_cell.fill = PatternFill("solid", fgColor="FFFFF7ED")
        warning_cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A1"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if model.width > model.height else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.sheet_view.zoomScale = 90


def _write_preview_sheet(ws, model: PageModel) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 140
    ws["A1"] = f"원본 PDF {model.number}페이지"
    ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True, color="FF172033")
    ws.row_dimensions[1].height = 24
    image = ExcelImage(io.BytesIO(model.preview_png))
    max_width = 900
    if image.width > max_width:
        ratio = max_width / image.width
        image.width = max_width
        image.height = round(image.height * ratio)
    ws.add_image(image, "A3")
    ws.sheet_view.zoomScale = 75


def build_workbook(models: list[PageModel], source_name: str) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = f"{source_name} PDF 변환"
    workbook.properties.creator = "KunhwaTools · pdfplumber"
    used_names: set[str] = set()
    for model in models:
        data_sheet = workbook.create_sheet(_safe_sheet_name(f"{model.number}페이지", used_names))
        _write_page_sheet(data_sheet, model)
        preview_sheet = workbook.create_sheet(_safe_sheet_name(f"{model.number}페이지_원본", used_names))
        _write_preview_sheet(preview_sheet, model)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    load_workbook(output, read_only=False, data_only=False).close()
    return output.getvalue()


def convert_pdf_bytes(pdf_bytes: bytes, source_name: str) -> tuple[bytes, dict]:
    if not pdf_bytes.startswith(b"%PDF-"):
        raise ValueError("PDF 파일 서명이 올바르지 않습니다.")
    models = analyze_pdf(pdf_bytes)
    workbook_bytes = build_workbook(models, source_name)
    report = {
        "engine": f"pdfplumber {pdfplumber.__version__}",
        "pages": len(models),
        "tables": sum(len(model.tables) for model in models),
        "text_lines": sum(len(model.lines) for model in models),
        "column_pages": sum(model.column_split is not None for model in models),
        "ocr_required_pages": sum(bool(model.warnings) for model in models),
        "sheet_count": len(models) * 2,
    }
    return workbook_bytes, report
